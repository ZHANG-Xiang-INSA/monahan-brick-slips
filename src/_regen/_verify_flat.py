#!/usr/bin/env python3
"""FULL verification for the FLAT-CORNER variant (run mid-way AND at the end)."""
import sys,types,json,collections,os,hashlib,math
import openpyxl,ezdxf
exec(open("_measure_corners.py").read().split("M=\"")[0])   # bpy stub + Vector
NR=".."
P=[]
def chk(lab,a,b): P.append((lab,a,b))
# ---------- 1. model ----------
M=NR+"/combined_red_black_brick_with_clips_red_d_se_garden_position_fix.py"
src=open(M).read().replace('    def add(s,base,normal,t):\n        k=len(base); off=Vector(normal).normalized()*t; b0=len(s.v)',
   '    def add(s,base,normal,t):\n        _BR.append((s.n,min(p.z for p in base),max(p.z for p in base)))\n        k=len(base); off=Vector(normal).normalized()*t; b0=len(s.v)',1)
g={"_BR":[],"__file__":os.path.abspath(M),"__name__":"x"}
exec(compile(src,M,"exec"),g)
BR=g["_BR"]; bc=collections.Counter(n for n,a,b in BR)
nred=sum(v for k,v in bc.items() if k.startswith("Red")); nblk=sum(v for k,v in bc.items() if k.startswith("Black"))
cnt=g["cnt"]
chk("model red L-slips == 0", bc.get("Red_L",0), 0)
chk("model corner flats == 234 (117 lap + 117 half)", (cnt["red_corner_lap"],cnt["red_corner_half"]), (117,117))
chk("model total == 7552", nred+nblk, 7552)
chk("model black == 1461 (unchanged)", nblk, 1461)
chk("model red == 6091", nred, 6091)
chk("model red slips below z=0 (basement stays removed)", sum(1 for n,a,b in BR if n.startswith("Red") and b<-0.001), 0)
# ---------- 2. brick_types ----------
BTD=json.load(open(NR+"/brick_types.json")); BT={t["type_id"]:t for t in BTD["types"]}
chk("model total == brick_types n_bricks", nred+nblk, BTD["n_bricks"])
chk("model total == sum(type qty)", nred+nblk, sum(m["qty"] for m in BT.values()))
chk("red L-corner type count == 0", sum(1 for m in BT.values() if m["colour"]=="Red" and m["kind"]=="L-corner"), 0)
chk("black L type qty == 272", sum(m["qty"] for m in BT.values() if m["colour"]=="Black" and m["kind"]=="L-corner"), 272)
# ---------- 3. per-wall DXF counts (count BRICK-layer polylines) ----------
tot=0; per={}
for w in ["A_NE_side","B_SW_side","C_NW_front","D_SE_garden"]:
    d=ezdxf.readfile(NR+"/brick_layout_wall_%s.dxf"%w)
    nn=sum(1 for e in d.modelspace() if e.dxftype()=="LWPOLYLINE" and e.dxf.layer=="BRICK")
    per[w]=nn; tot+=nn
dblk=ezdxf.readfile(NR+"/brick_layout_black_canopy_faces.dxf")
nb=sum(1 for e in dblk.modelspace() if e.dxftype()=="LWPOLYLINE" and e.dxf.layer=="BRICK")
chk("per-wall DXF red sum == 6091", tot, 6091)
chk("black-face DXF bricks == 1461", nb, 1461)
chk("model total == per-wall DXF sum", nred+nblk, tot+nb)
# brick cutting DXF headers == n types
d=ezdxf.readfile(NR+"/brick_tile_cutting_layout_red_d_se_garden_position_fix.dxf")
nh=sum(1 for e in d.modelspace() if e.dxftype()=="TEXT" and e.dxf.text.split()[:1] and e.dxf.text.split()[0].startswith("B") and e.dxf.text.split()[0][1:].isdigit())
chk("brick cutting DXF type cells == %d"%len(BT), nh, len(BT))
# brick Excel
wb=openpyxl.load_workbook(NR+"/brick_tile_schedule_detailed_red_d_se_garden_position_fix.xlsx",data_only=True)
ws=wb["Brick Procurement Summary"];rows=list(ws.iter_rows(values_only=True))
hd=[i for i,r in enumerate(rows) if r and any(isinstance(c,str) and c.strip() in("Type ID","Brick Type ID") for c in r)][0]
tc=next(j for j,c in enumerate(rows[hd]) if isinstance(c,str) and "Type" in c);qc=next(j for j,c in enumerate(rows[hd]) if isinstance(c,str) and c.strip()=="Quantity")
exq={r[tc].strip():r[qc] for r in rows[hd+1:] if r and isinstance(r[tc],str) and r[tc].strip().startswith("B") and isinstance(r[qc],int)}
chk("brick Excel Procurement total == 7552", sum(exq.values()), 7552)
chk("brick Excel per-type == brick_types (mismatches)", sum(1 for k in BT if exq.get(k)!=BT[k]["qty"]), 0)
# ---------- 4. clips ----------
CI=json.load(open(NR+"/clip_instances_red_d_se_garden_position_fix.json"));clips=CI["clips"]
mclip=collections.Counter(c["type_id"] for c in clips)
dxfq={t["type_id"]:t["qty"] for t in CI["types"]}
wb2=openpyxl.load_workbook(NR+"/clip_schedule_red_d_se_garden_position_fix.xlsx",data_only=True);ws2=wb2["Procurement Summary"];r2=list(ws2.iter_rows(values_only=True))
h2=[i for i,r in enumerate(r2) if r and any(isinstance(c,str) and "quantity" in c.lower() for c in r)][0]
t2=next(j for j,c in enumerate(r2[h2]) if isinstance(c,str) and "type" in c.lower());q2=next(j for j,c in enumerate(r2[h2]) if isinstance(c,str) and "quantity" in c.lower())
xl2={r[t2].strip():r[q2] for r in r2[h2+1:] if r and isinstance(r[t2],str) and r[t2][:1] in "RT" and isinstance(r[q2],int)}
nmodelclips=len(g.get("CLIPS",[]))
chk("with-clips model loads clip json (count)", nmodelclips, len(clips))
chk("clip totals model==excel==dxf", (len(clips),sum(xl2.values()),sum(dxfq.values())), (len(clips),)*3)
chk("clip per-type model==excel==dxf", all(mclip.get(t)==xl2.get(t)==dxfq.get(t) for t in set(list(mclip)+list(xl2)+list(dxfq))), True)
# cutting DXF: count type header texts
d=ezdxf.readfile(NR+"/clip_cutting_layout_red_d_se_garden_position_fix.dxf")
nhc=sum(1 for e in d.modelspace() if e.dxftype()=="TEXT" and (e.dxf.text.split()[:1] and (e.dxf.text.split()[0].startswith("R") or e.dxf.text.split()[0].startswith("T_")) and "[" in e.dxf.text))
chk("clip cutting DXF cells == types in json", nhc, len(CI["types"]))
# red clips: on-structure + on-brick (coverage incl corner flats, clamped)
RED=json.load(open(NR+"/red_brick_placement_v7_stairfix.json"))
MMc=0.001; WXc=11513.0; WYc=13604.0
RFRAME={"C_NW_front":((0,0,0),(1,0,0)),"A_NE_side":((0,WYc*MMc,0),(0,-1,0)),
        "D_SE_garden":(((WXc+997)*MMc,WYc*MMc,0),(-1,0,0)),"B_SW_side":((WXc*MMc,0,0),(0,1,0))}
UDOM={"C_NW_front":(0.0,11513.0),"A_NE_side":(0.0,13604.0),"B_SW_side":(0.0,None),"D_SE_garden":(None,12510.0)}
def runs_of(wall):
    cs=collections.defaultdict(list)
    ulo,uhi=UDOM[wall]
    for br in RED[wall]["bricks"]:
        us=[v[0] for v in br["verts"]]; zs=[v[1] for v in br["verts"]]
        u0,u1=min(us),max(us)
        if ulo is not None: u0=max(u0,ulo)
        if uhi is not None: u1=min(u1,uhi)
        if u1-u0<=1: continue
        cs[round(((min(zs)+max(zs))/2)/75.0)].append((u0,u1))
    R={}
    for k,iv in cs.items():
        iv.sort(); out=[]; a0,a1=iv[0]
        for b0,b1 in iv[1:]:
            if b0-a1<=16.0: a1=max(a1,b1)
            else: out.append((a0,a1)); a0,a1=b0,b1
        out.append((a0,a1)); R[k]=out
    return R
RUNS={w:runs_of(w) for w in RFRAME}
bad_float=0; bad_dom=0
for c in clips:
    if c["colour"]!="Red": continue
    w=c["face"]; O,U=RFRAME[w]
    us=[]; vs=[]
    for pxyz in c["poly"]:
        rel=[pxyz[i]-O[i] for i in range(3)]
        us.append((rel[0]*U[0]+rel[1]*U[1]+rel[2]*U[2])/MMc); vs.append(pxyz[2]/MMc)
    u0,u1=min(us),max(us); vc=(min(vs)+max(vs))/2
    ulo,uhi=UDOM[w]
    if (ulo is not None and u0<ulo-0.01) or (uhi is not None and u1>uhi+0.01): bad_dom+=1
    k=round(vc/75.0-0.5+0.5)  # nearest band
    kk=int(round((vc-32.5)/75.0)) if False else min(RUNS[w].keys(), key=lambda q: abs(q*75+32.5-vc))
    ok=any(u0>=a0-0.6 and u1<=a1+0.6 for (a0,a1) in RUNS[w][kk])
    if not ok: bad_float+=1
chk("red clips outside structural wall domain", bad_dom, 0)
chk("red clips floating / overhanging brick coverage", bad_float, 0)
zbad=sum(1 for c in clips if any(len(v)>=3 and v[2]<-0.001 for v in c["poly"]))
chk("clips below grade", zbad, 0)
# black clips byte-identical vs backup
OLDCI=json.load(open("_clips_PREFLAT.bak.json"))
def kk2(c): return (c["colour"],c["face"],c["kind"],c["length"],c["source"],tuple(tuple(p) for p in c["poly"]))
chk("black clips identical to pre-change", sorted(map(kk2,[c for c in OLDCI["clips"] if c["colour"]=="Black"]))==sorted(map(kk2,[c for c in clips if c["colour"]=="Black"])), True)
# ---------- 5. BLACK untouched ----------
md5=hashlib.md5(open(NR+"/black_placement_fixed7.json","rb").read()).hexdigest()
rec=open("_black_md5_PREFLAT.txt").read().split()[0]
chk("black_placement_fixed7.json md5 unchanged", md5, rec)
OLDBT={t["type_id"]:t for t in json.load(open("_brick_types_PREFLAT.bak.json"))["types"]}
oldblk={k:(v["qty"],v["w"],v["h"],v["kind"]) for k,v in OLDBT.items() if v["colour"]=="Black"}
newblk={k:(v["qty"],v["w"],v["h"],v["kind"]) for k,v in BT.items() if v["colour"]=="Black"}
chk("black brick types identical (id/qty/size/kind)", oldblk==newblk, True)
chk("black total area unchanged", BTD["black_area"], json.load(open("_brick_types_PREFLAT.bak.json"))["black_area"])
# ---------- 6. field unchanged except re-anchored ----------
OLDP=json.load(open(NR+"/red_brick_placement_v7_stairfix_PREFLAT.bak.json"))
def canon(R,w,field_only=True):
    s=set()
    for b in R[w]["bricks"]:
        if field_only and b.get("cp"): continue
        s.add((b["t"],tuple(sorted((round(v[0],3),round(v[1],3)) for v in b["verts"]))))
    return s
tots=0; totc=0
for w in RFRAME:
    o=canon(OLDP,w,False); n=canon(RED,w,True)
    tots+=len(o&n); totc+=len(o-n)
chk("old field bricks disturbed == 468 (234 b1 + 234 b2)", totc, 468)
chk("field bricks byte-identical == 5272", tots, 5272)
nmark=collections.Counter()
for w in RFRAME:
    for b in RED[w]["bricks"]:
        if b.get("cp"): nmark[b["cp"]]+=1
chk("markers lap/half/pair/trim == 117/117/234/117", (nmark["lap"],nmark["half"],nmark["pair"],nmark["trim"]), (117,117,234,117))
# ---------- 7. corner geometry (delegated audit) ----------
import subprocess
r=subprocess.run([sys.executable,"_audit_corners_flat.py"],capture_output=True,text=True)
aud=r.stdout.strip().splitlines()
chk("corner audit: overlaps 0 / gaps 0", aud[0].startswith("overlaps: 0 | suspicious gaps: 0"), True)
chk("corner audit: 234 joints all 10mm", any("out of tolerance: 0" in l for l in aud), True)
chk("corner audit: FAILS 0", any(l.startswith("FAILS: 0") for l in aud), True)
ok=all(a==b for _,a,b in P)
print()
for lab,a,b in P: print(("PASS  " if a==b else "FAIL  ")+lab+": %s vs %s"%(a,b))
print("\nRESULT:", "ALL PASS" if ok else "HAS FAILURES")
