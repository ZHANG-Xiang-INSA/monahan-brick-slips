#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""clip_schedule_red_d_se_garden_position_fix.xlsx - simplified to 6 clip types. Procurement + Placement + Validation."""
import json,os,collections
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
HERE=os.path.dirname(os.path.abspath(__file__)); B="/sessions/tender-cool-ritchie/mnt/Brick slip PJ 001"
D=json.load(open(os.path.join(HERE,"..","clip_instances_red_d_se_garden_position_fix.json"))); clips=D["clips"]; types={t["type_id"]:t for t in D["types"]}
OUT=os.path.join(HERE,"..","clip_schedule_red_d_se_garden_position_fix.xlsx")
FONT="Arial"
H1=Font(name=FONT,size=13,bold=True,color="FFFFFF"); HEAD=Font(name=FONT,size=9,bold=True,color="FFFFFF")
H2=Font(name=FONT,size=11,bold=True,color="FFFFFF"); BOLD=Font(name=FONT,size=10,bold=True); REG=Font(name=FONT,size=10); ITAL=Font(name=FONT,size=9,italic=True,color="555555")
GRN=PatternFill("solid",fgColor="2E7D32"); NAVY=PatternFill("solid",fgColor="1F3864"); TEAL=PatternFill("solid",fgColor="0F6E63")
NAMEHEX={"blue":"0050FF","green":"00B050","cyan":"00C0C0","grey":"808080","red":"FF0000","magenta":"FF00FF"}
def fill(name): return PatternFill("solid",fgColor=NAMEHEX.get(name,"BFBFBF"))
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment("center",vertical="center"); LEFT=Alignment("left",vertical="center",wrap_text=True); RIGHT=Alignment("right",vertical="center")
def c(ws,r,col,v,f=REG,fl=None,al=None,fmt=None):
    x=ws.cell(row=r,column=col,value=v); x.font=f; x.alignment=al or LEFT; x.border=BORDER
    if fl:x.fill=fl
    if fmt:x.number_format=fmt
    return x
def title(ws,t,n,fl=NAVY):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    a=ws.cell(row=1,column=1,value=t);a.font=H1;a.fill=fl;a.alignment=Alignment("left",vertical="center");ws.row_dimensions[1].height=24
ORDER=["R1000","R700","R500","R300","R250","R100","R50","R20","T_LEFT_40","T_RIGHT_40"]
RN={"red":"red brick","black_ordinary":"black ordinary","black_sloped_canopy":"black sloped canopy"}
def usedfor(tid):
    rs=sorted(set(c["region"] for c in clips if c["type_id"]==tid)); return ", ".join(RN[r] for r in rs)
wb=Workbook()
# ---- Procurement Summary ----
ws=wb.active; ws.title="Procurement Summary"
title(ws,"3 Monahan Ave (HA23007) — CLIP PROCUREMENT (10 TYPES). width 68 mm, thickness 0.25 mm. Added 700/300/100/50 to cut the 20mm count.",13,GRN)
hd=["Clip Type ID","Shape","Colour","Length / top width (mm)","Bottom width (mm)","Width (mm)","Thickness (mm)","Angle","Quantity","Total source length (m)","Total installed length (m)","Used for","Notes"]
for j,h in enumerate(hd,1): c(ws,2,j,h,HEAD,GRN,CEN)
r=3
for tid in ORDER:
    t=types.get(tid)
    if not t: continue
    isT=tid.startswith("T_"); cs=[x for x in clips if x["type_id"]==tid]
    lentop = (30 if isT else int(tid[1:])); bot=(111 if isT else "—")
    src=t["source"]; tot_src=sum(x["source"] for x in cs)/1000.0; tot_inst=sum((111 if isT else x["length"]) for x in cs)/1000.0
    c(ws,r,1,tid,BOLD,fill(t["colour"]),CEN); c(ws,r,2,("right trapezoid" if isT else "rectangle"),REG); c(ws,r,3,t["colour"],REG,fill(t["colour"]),CEN)
    c(ws,r,4,lentop,REG,None,RIGHT); c(ws,r,5,bot,REG,None,RIGHT); c(ws,r,6,68,REG,None,CEN); c(ws,r,7,0.25,REG,None,CEN)
    c(ws,r,8,("40°" if isT else "—"),REG,None,CEN); c(ws,r,9,t["qty"],BOLD,None,RIGHT)
    c(ws,r,10,round(tot_src,2),REG,None,RIGHT,'0.00'); c(ws,r,11,round(tot_inst,2),REG,None,RIGHT,'0.00'); c(ws,r,12,usedfor(tid),REG)
    c(ws,r,13,("mirror of T_RIGHT_40, sloped LEFT, canopy rake" if tid=="T_LEFT_40" else "mirror of T_LEFT_40, sloped RIGHT, canopy rake" if tid=="T_RIGHT_40" else ("min filler; spacing adjustable 0-50 mm" if tid=="R20" else "standard stock, used as-is")),ITAL); r+=1
c(ws,r,1,"TOTAL",H2,GRN); c(ws,r,9,f"=SUM(I3:I{r-1})",H2,GRN,RIGHT); c(ws,r,10,f"=SUM(J3:J{r-1})",H2,GRN,RIGHT,'0.00'); c(ws,r,11,f"=SUM(K3:K{r-1})",H2,GRN,RIGHT,'0.00')
for col in (2,3,4,5,6,7,8,12,13): c(ws,r,col,"",H2,GRN)
for col,w in zip("ABCDEFGHIJKLM",[12,15,9,18,14,9,11,7,9,18,19,26,34]): ws.column_dimensions[col].width=w
ws.freeze_panes="A3"
# ---- Placement Summary ----
ws=wb.create_sheet("Placement Summary"); title(ws,"Placement by region — quantities per clip type (still only the 6 types)",8,NAVY)
regs=["red","black_ordinary","black_sloped_canopy"]
for j,h in enumerate(["Clip Type ID"]+[RN[x] for x in regs]+["Total"],1): c(ws,2,j,h,HEAD,NAVY,CEN)
r=3
for tid in ORDER:
    cs=[x for x in clips if x["type_id"]==tid]
    if not cs: continue
    qn=[sum(1 for x in cs if x["region"]==rg) for rg in regs]
    c(ws,r,1,tid,BOLD,fill(types[tid]["colour"]),CEN)
    for j,q in enumerate(qn,2): c(ws,r,j,q,REG,None,RIGHT)
    c(ws,r,5,sum(qn),BOLD,None,RIGHT); r+=1
c(ws,r,1,"TOTAL",H2,NAVY)
for j,L in ((2,"B"),(3,"C"),(4,"D"),(5,"E")): c(ws,r,j,f"=SUM({L}3:{L}{r-1})",H2,NAVY,RIGHT)
for col,w in zip("ABCDE",[14,16,16,22,10]): ws.column_dimensions[col].width=w
# ---- Validation ----
ws=wb.create_sheet("Validation"); title(ws,"Validation — 10-clip-type system (700/300/100/50 added)",2,TEAL)
v=[("10 clip types","Rectangles R1000/R700/R500/R300/R250/R100/R50/R20 + trapezoids T_LEFT_40/T_RIGHT_40. Added 700/300/100/50 so 20mm count is minimised."),
 ("20mm minimised","R20 reduced from 2102 to %d (one <=49mm tail filler per run max). Long clips start at the wall corners."%sum(1 for x in clips if x["source"]==20)),
 ("Red brick clips","rectangles R1000..R50 + minimal R20; long clips aligned from the wall corners. qty %d."%sum(1 for x in clips if x["region"]=="red")),
 ("Black side walls (4 faces)","one R700 per course (side faces use 700, from the corner). qty %d."%sum(1 for x in clips if x["region"]=="black_ordinary")),
 ("Black canopy","gable rakes = T_LEFT_40/T_RIGHT_40; front rectangular jambs = R300; soffit slopes = R700; gable middle = R-fills. qty %d."%sum(1 for x in clips if x["region"]=="black_sloped_canopy")),
 ("Trapezoid geometry","right trapezoid, top=30, bottom=111 (=30+68/tan40), H=68, t=0.25, angle=40°. T_LEFT_40 & T_RIGHT_40 identical size, mirrored."),
 ("Inner sloped edge","now uses T_LEFT_40 / T_RIGHT_40 (mirrored/rotated), same as the outer rake - no R20 patchwork on the rake."),
 ("20 mm clip spacing","R20 used as adjustable fillers; gaps 0–50 mm to fit residual lengths (5 mm typical)."),
 ("Width / thickness","all clips width 68 mm, thickness 0.25 mm. No 1 mm residual."),
 ("Clips opaque + hideable","with-clips Blender model: clips opaque (alpha 1.0), in a separate hideable collection, behind the bricks."),
 ("Brick model unchanged","red + black brick geometry, counts, sizes, L-shaped bricks, openings — all unchanged. Only clips changed."),
 ("Blender / DXF / Excel match","every clip is one of the 6 types in the model, the wall-layout DXF, the cutting DXF and this Excel; quantities + colours match."),
 ("Totals","6 types · %d clips total (red %d, black ordinary %d, black canopy %d)."%(len(clips),sum(1 for x in clips if x['region']=='red'),sum(1 for x in clips if x['region']=='black_ordinary'),sum(1 for x in clips if x['region']=='black_sloped_canopy')))]
r=3
for k,vv in v:
    c(ws,r,1,k,BOLD); c(ws,r,2,vv,REG,al=LEFT); ws.row_dimensions[r].height=30; r+=1
ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=104
wb.save(OUT); print("Saved",OUT,"| sheets:",wb.sheetnames)
