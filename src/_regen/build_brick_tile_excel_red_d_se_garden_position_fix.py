#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""brick_tile_schedule_detailed_red_d_se_garden_position_fix.xlsx — full brick-slip schedule (NOT clips). 6 sheets,
area-based, Red/Black/Cut/L separated. Source = brick_types.json (the current Blender model)."""
import json,os,collections
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
HERE=os.path.dirname(os.path.abspath(__file__))
D=json.load(open(os.path.join(HERE,"..","brick_types.json"))); T=D["types"]
OUT=os.path.join(HERE,"..","brick_tile_schedule_detailed_red_d_se_garden_position_fix.xlsx")
FONT="Arial"
H1=Font(name=FONT,size=14,bold=True,color="FFFFFF"); HEAD=Font(name=FONT,size=9,bold=True,color="FFFFFF")
H2=Font(name=FONT,size=11,bold=True,color="FFFFFF"); BOLD=Font(name=FONT,size=10,bold=True); REG=Font(name=FONT,size=10)
ITAL=Font(name=FONT,size=9,italic=True,color="555555")
NAVY=PatternFill("solid",fgColor="1F3864"); GRN=PatternFill("solid",fgColor="2E7D32"); REDF=PatternFill("solid",fgColor="9E2B0E")
BLKF=PatternFill("solid",fgColor="2E2E38"); TEAL=PatternFill("solid",fgColor="0F6E63"); ORG=PatternFill("solid",fgColor="B5651D")
RTINT=PatternFill("solid",fgColor="FCE4D6"); BTINT=PatternFill("solid",fgColor="E2E2E8")
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment("center",vertical="center"); LEFT=Alignment("left",vertical="center",wrap_text=True); RIGHT=Alignment("right",vertical="center")
def c(ws,r,col,v,f=REG,fill=None,al=None,fmt=None):
    x=ws.cell(row=r,column=col,value=v); x.font=f; x.alignment=al or LEFT; x.border=BORDER
    if fill:x.fill=fill
    if fmt:x.number_format=fmt
    return x
def title(ws,t,n,fill=NAVY):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    a=ws.cell(row=1,column=1,value=t);a.font=H1;a.fill=fill;a.alignment=Alignment("left",vertical="center");ws.row_dimensions[1].height=24
def cutreason(t):
    if t["kind"]=="L-corner": return "corner (L-shaped wrap)"
    if t["colour"]=="Black":
        if t["kind"]=="angled": return "sloped canopy (40 deg rake / gable / apex)"
        return "canopy gable / slope edge / opening"
    return "wall edge / window-door opening / stair edge"
def rng(t):
    sr=t.get("size_range")
    if not sr: return str(t["w"]),str(t["h"])
    wt=f"{sr[0]}-{sr[1]}" if sr[0]!=sr[1] else str(t["w"]); ht=f"{sr[2]}-{sr[3]}" if sr[2]!=sr[3] else str(t["h"]); return wt,ht
def dims_txt(t):
    wt,ht=rng(t)
    if t["shape"]=="L-shaped corner": return f"{t['edges'][0]} + {t['edges'][1]} x {t['edges'][2]} (stretcher+header x H)"
    if t["kind"]=="angled": return f"{wt} x {ht} bbox; edges {('/'.join(str(e) for e in t['edges']))}"
    return f"{wt} x {ht}"
wb=Workbook()

# ---------- Sheet 1: Brick Procurement Summary ----------
ws=wb.active; ws.title="Brick Procurement Summary"
title(ws,"3 Monahan Ave (HA23007) — BRICK SLIP PROCUREMENT SUMMARY (slips, NOT clips). 10 mm joint. Areas in m2.",10,GRN)
hd=["Brick Type ID","Colour / material","Shape type","Standard/Cut/L","Dimensions (mm)","Cut angle","Quantity","Area / piece (mm2)","Total area (m2)","Notes / source"]
for j,h in enumerate(hd,1): c(ws,2,j,h,HEAD,GRN,CEN)
rr=3
for t in T:
    fill=RTINT if t["colour"]=="Red" else BTINT
    c(ws,rr,1,t["type_id"],BOLD,fill,CEN); c(ws,rr,2,t["colour"]+" brick slip",REG,fill); c(ws,rr,3,t["shape"],REG)
    c(ws,rr,4,t["kind"],REG,None,CEN); c(ws,rr,5,dims_txt(t),REG); c(ws,rr,6,(f"{t['angle']:.0f} deg" if t["angle"]>=5 else "—"),REG,None,CEN)
    c(ws,rr,7,t["qty"],BOLD,None,RIGHT); c(ws,rr,8,t["area_per"],REG,None,RIGHT); c(ws,rr,9,round(t["total_area"]/1e6,3),REG,None,RIGHT,'0.000')
    c(ws,rr,10,", ".join(t["regions"]),ITAL); rr+=1
c(ws,rr,1,"TOTAL",H2,GRN); c(ws,rr,7,f"=SUM(G3:G{rr-1})",H2,GRN,RIGHT); c(ws,rr,9,f"=SUM(I3:I{rr-1})",H2,GRN,RIGHT,'0.000')
for col in (2,3,4,5,6,8,10): c(ws,rr,col,"",H2,GRN)
for col,w in zip("ABCDEFGHIJ",[12,16,16,12,30,9,9,15,13,34]): ws.column_dimensions[col].width=w
ws.freeze_panes="A3"

# ---------- Sheets 2 & 3: Red / Black tile schedule (per type, per region) ----------
def colour_sheet(name,colour,fillhdr):
    ws=wb.create_sheet(name); title(ws,f"{colour} brick slip schedule — by Type ID and wall/component. Areas in m2.",11,fillhdr)
    hd=["Brick Type ID","Wall / component","Shape type","Length (mm)","Height (mm)","Cut angle","Area/pc (mm2)","Qty (this region)","Qty (total)","Total area (m2)","Notes"]
    for j,h in enumerate(hd,1): c(ws,2,j,h,HEAD,fillhdr,CEN)
    rr=3
    for t in [x for x in T if x["colour"]==colour]:
        for reg,q in sorted(t["region_qty"].items()):
            c(ws,rr,1,t["type_id"],BOLD,None,CEN); c(ws,rr,2,reg,REG); c(ws,rr,3,t["shape"],REG)
            c(ws,rr,4,(t["w"] if t["shape"]!="L-shaped corner" else f"{t['edges'][0]}+{t['edges'][1]}"),REG,None,RIGHT)
            c(ws,rr,5,t["h"],REG,None,RIGHT); c(ws,rr,6,(f"{t['angle']:.0f}" if t["angle"]>=5 else "—"),REG,None,CEN)
            c(ws,rr,7,t["area_per"],REG,None,RIGHT); c(ws,rr,8,q,REG,None,RIGHT); c(ws,rr,9,t["qty"],REG,None,RIGHT)
            c(ws,rr,10,round(t["area_per"]*q/1e6,3),REG,None,RIGHT,'0.000'); c(ws,rr,11,t["kind"],ITAL); rr+=1
    c(ws,rr,1,"TOTAL",H2,fillhdr); c(ws,rr,8,f"=SUM(H3:H{rr-1})",H2,fillhdr,RIGHT); c(ws,rr,10,f"=SUM(J3:J{rr-1})",H2,fillhdr,RIGHT,'0.000')
    for col in (2,3,4,5,6,7,9,11): c(ws,rr,col,"",H2,fillhdr)
    for col,w in zip("ABCDEFGHIJK",[12,26,16,11,9,8,12,14,11,13,12]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A3"
colour_sheet("Red Brick Tile Schedule","Red",REDF)
colour_sheet("Black Brick Tile Schedule","Black",BLKF)

# ---------- Sheet 4: Cut Brick Schedule ----------
ws=wb.create_sheet("Cut Brick Schedule"); title(ws,"CUT brick slips (non-full rectangles + angled cuts). Edge lengths, angle, reason. Areas in m2.",11,ORG)
hd=["Brick Type ID","Colour","Wall / component","Shape type","Edge lengths (mm)","Cut angle","Quantity","Area/pc (mm2)","Total area (m2)","Reason for cut","Notes"]
for j,h in enumerate(hd,1): c(ws,2,j,h,HEAD,ORG,CEN)
rr=3
for t in [x for x in T if x["kind"] in ("cut","angled")]:
    fill=RTINT if t["colour"]=="Red" else BTINT
    c(ws,rr,1,t["type_id"],BOLD,fill,CEN); c(ws,rr,2,t["colour"],REG,fill,CEN); c(ws,rr,3,", ".join(t["regions"]),REG)
    c(ws,rr,4,t["shape"],REG); c(ws,rr,5,"/".join(str(e) for e in t["edges"]) if t["kind"]=="angled" else f"{t['w']} x {t['h']}",REG)
    c(ws,rr,6,(f"{t['angle']:.0f} deg" if t["angle"]>=5 else "square"),REG,None,CEN); c(ws,rr,7,t["qty"],BOLD,None,RIGHT)
    c(ws,rr,8,t["area_per"],REG,None,RIGHT); c(ws,rr,9,round(t["total_area"]/1e6,3),REG,None,RIGHT,'0.000'); c(ws,rr,10,cutreason(t),REG); c(ws,rr,11,t["kind"],ITAL); rr+=1
c(ws,rr,1,"TOTAL cut",H2,ORG); c(ws,rr,7,f"=SUM(G3:G{rr-1})",H2,ORG,RIGHT); c(ws,rr,9,f"=SUM(I3:I{rr-1})",H2,ORG,RIGHT,'0.000')
for col in (2,3,4,5,6,8,10,11): c(ws,rr,col,"",H2,ORG)
for col,w in zip("ABCDEFGHIJK",[12,8,24,16,20,9,9,13,13,30,9]): ws.column_dimensions[col].width=w
ws.freeze_panes="A3"

# ---------- Sheet 5: L-shaped Brick Schedule ----------
ws=wb.create_sheet("L-shaped Brick Schedule"); title(ws,"L-SHAPED corner brick slips. Stretcher + header faces, height 65. Area = unfolded two-face. m2.",8,TEAL)
hd=["Brick Type ID","Colour","Corner / location","Stretcher (mm)","Header / return (mm)","Height (mm)","Quantity","Area/pc (mm2)","Total area (m2)","Notes"]
for j,h in enumerate(hd,1): c(ws,2,j,h,HEAD,TEAL,CEN)
rr=3
for t in [x for x in T if x["kind"]=="L-corner"]:
    for reg,q in sorted(t["region_qty"].items()):
        c(ws,rr,1,t["type_id"],BOLD,None,CEN); c(ws,rr,2,t["colour"],REG,None,CEN); c(ws,rr,3,reg,REG)
        c(ws,rr,4,t["edges"][0],REG,None,RIGHT); c(ws,rr,5,t["edges"][1],REG,None,RIGHT); c(ws,rr,6,t["edges"][2],REG,None,RIGHT)
        c(ws,rr,7,q,BOLD,None,RIGHT); c(ws,rr,8,t["area_per"],REG,None,RIGHT); c(ws,rr,9,round(t["area_per"]*q/1e6,3),REG,None,RIGHT,'0.000')
        c(ws,rr,10,"one-piece L wraps the corner (two visible faces)",ITAL); rr+=1
if rr>3:
    c(ws,rr,1,"TOTAL L",H2,TEAL); c(ws,rr,7,f"=SUM(G3:G{rr-1})",H2,TEAL,RIGHT); c(ws,rr,9,f"=SUM(I3:I{rr-1})",H2,TEAL,RIGHT,'0.000')
else:
    c(ws,rr,1,"None - all corners are flat slip pairs; no L-shaped slips",H2,TEAL); c(ws,rr,7,0,H2,TEAL,RIGHT); c(ws,rr,9,0.0,H2,TEAL,RIGHT,'0.000')
for col in (2,3,4,5,6,8,10): c(ws,rr,col,"",H2,TEAL)
for col,w in zip("ABCDEFGHIJ",[12,8,26,13,16,10,9,13,13,40]): ws.column_dimensions[col].width=w

# ---------- Sheet 6: Assumptions & Validation ----------
ws=wb.create_sheet("Assumptions & Validation"); title(ws,"Brick-slip schedule — assumptions, area method, validation",2)
ra=D["red_area"]/1e6; ba=D["black_area"]/1e6
# area by shape
byshape=collections.defaultdict(float)
for t in T: byshape[t["shape"]]+=t["total_area"]/1e6
items=[("Source","Current Blender model placements: red_brick_placement_v7_stairfix.json + black_placement_fixed7.json (same files the with-clips model is built from)."),
 ("Mortar joint","10 mm (already in the placement geometry)."),
 ("Black canopy top depth","700 mm (inner soffit slope; = black vertical elevation depth). Checked, not changed."),
 ("Brick face","Standard slip face 215 x 65 mm; thickness 20 mm (slip)."),
 ("Tile grouping tolerance","Same SIZE = one Brick Type ID. Bricks are grouped by colour + shape + angle, then clustered by length and height within +-5 mm (the per-type size range is shown). STANDARD = 215 x 65 +-3 mm (full tile); anything else is a CUT. Standard vs cut is decided by SIZE, never split into two IDs for the same size."),
 ("Area per piece","Rectangles: L x H. Angled (trapezoid/parallelogram/triangle): true polygon area. L-corner: unfolded two visible faces = (stretcher + header) x height."),
 ("L-shaped area method","Unfolded face area of the two visible legs (NOT footprint). Stated per row."),
 ("Total Red brick area","%.2f m2"%ra),
 ("Total Black brick area","%.2f m2"%ba),
 ("Total brick area","%.2f m2"%(ra+ba)),
 ("Area by shape (m2)",", ".join(f"{k} {v:.2f}" for k,v in sorted(byshape.items(),key=lambda x:-x[1]))),
 ("Brick Type IDs","%d total (Red %d, Black %d). Standard 2, cut %d, angled %d, L-corner 1."%(len(T),
     sum(1 for x in T if x['colour']=='Red'),sum(1 for x in T if x['colour']=='Black'),
     sum(1 for x in T if x['kind']=='cut'),sum(1 for x in T if x['kind']=='angled'))),
 ("DXF correspondence","brick_tile_cutting_layout.dxf draws the SAME Brick Type IDs / shapes / dims / qty; one cell per type."),
 ("Red vs Black","Separated in every sheet and in the DXF (swatch colour)."),
 ("Cut vs L vs standard","Separate sheets (4 & 5) and DXF sections."),
 ("Manual review","Tiny qty-1 unique cuts are kept as their own Type ID (not merged) so nothing is lost.")]
rr=3
for k,v in items:
    c(ws,rr,1,k,BOLD); c(ws,rr,2,v,REG,al=LEFT); ws.row_dimensions[rr].height=30; rr+=1
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=110
wb.save(OUT); print("Saved",OUT); print("Sheets:",wb.sheetnames)
